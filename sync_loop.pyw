"""
sync_loop.pyw — sync setiap 30 detik, workbook di-cache di memory.

Perubahan dari versi sebelumnya:
- Tidak pakai subprocess untuk sync (menghindari timeout 25s yang membunuh proses
  sebelum Excel selesai dibuka — Excel butuh ~64 detik untuk load file besar)
- Workbook di-load SEKALI saat startup, lalu dipakai terus (write langsung ke cache)
- Hanya save + export + push kalau ada transaksi baru
"""
import sys
import os
import time
import subprocess
import ssl
import json
import urllib.request
import openpyxl
from datetime import datetime

# ── Setup ──────────────────────────────────────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

from dotenv import load_dotenv
load_dotenv(os.path.join(script_dir, ".env"))

RAILWAY_URL = os.getenv("RAILWAY_URL", "").rstrip("/")
SYNC_SECRET = os.getenv("SYNC_SECRET", "")
EXCEL_PATH  = os.getenv("EXCEL_PATH", r"C:\Claude\Project\Financial\Financial Management.xlsx")
INTERVAL    = 30  # detik antar sync

MONTH_SHEETS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
NTD_FMT = '_-"NT$ "* #,##0.00_-;\\-"NT$ "* #,##0.00_-;_-"NT$ "* "-"??_-;_-@_-'
NO_WIN  = subprocess.CREATE_NO_WINDOW
_SSL    = ssl._create_unverified_context()

# ── Workbook cache ─────────────────────────────────────────────────────────────
_wb = None

def get_wb():
    global _wb
    if _wb is None:
        _wb = openpyxl.load_workbook(EXCEL_PATH)
    return _wb

def invalidate_wb():
    global _wb
    _wb = None

# ── HTTP helpers ───────────────────────────────────────────────────────────────
def api_get(path):
    try:
        url = f"{RAILWAY_URL}{path}?token={SYNC_SECRET}"
        with urllib.request.urlopen(url, timeout=10, context=_SSL) as r:
            return json.loads(r.read())
    except Exception:
        return None

def api_post(path, data):
    try:
        url  = f"{RAILWAY_URL}{path}?token={SYNC_SECRET}"
        body = json.dumps(data).encode()
        req  = urllib.request.Request(url, data=body,
                                      headers={"Content-Type": "application/json"},
                                      method="POST")
        with urllib.request.urlopen(req, timeout=10, context=_SSL) as r:
            return json.loads(r.read())
    except Exception:
        return None

# ── Excel writer ───────────────────────────────────────────────────────────────
def find_next_row(ws):
    for row in range(11, 1001):  # row 10 = header, jangan ditimpa
        if ws.cell(row=row, column=2).value is None:
            return row
    return None

def do_sync():
    if not RAILWAY_URL or not SYNC_SECRET:
        return

    pending = api_get("/pending")
    if not pending:
        return

    try:
        wb = get_wb()
    except PermissionError:
        return
    except Exception:
        invalidate_wb()
        return

    written, done_ids = 0, []

    for tx in pending:
        try:
            date_obj   = datetime.strptime(tx["date"], "%Y-%m-%d").date()
            sheet_name = MONTH_SHEETS[date_obj.month - 1]
        except Exception:
            continue

        if sheet_name not in wb.sheetnames:
            continue

        ws  = wb[sheet_name]
        row = find_next_row(ws)
        if row is None:
            continue

        col_amount = 9 if tx.get("currency", "IDR") == "IDR" else 21
        ws.cell(row=row, column=2).value         = date_obj
        ws.cell(row=row, column=3).value         = tx["type"]
        ws.cell(row=row, column=6).value         = tx["category"]
        ws.cell(row=row, column=col_amount).value = float(tx["amount"])
        ws.cell(row=row, column=11).value        = tx["description"]
        ws.cell(row=row, column=2).number_format = "DD/MM/YYYY"
        if tx.get("currency", "IDR") == "NTD":
            ws.cell(row=row, column=21).number_format = NTD_FMT

        written  += 1
        done_ids.append(tx["id"])

    if written > 0:
        try:
            wb.save(EXCEL_PATH)
        except PermissionError:
            invalidate_wb()
            return
        api_post("/mark_synced", {"ids": done_ids})

    return written

# ── Export + push ──────────────────────────────────────────────────────────────
def do_export():
    try:
        from export_json import export
        export()
    except Exception:
        pass

def run_git(cmd):
    try:
        subprocess.run(cmd, capture_output=True, creationflags=NO_WIN, timeout=45)
    except Exception:
        pass

def push_if_changed():
    try:
        r = subprocess.run(
            ["git", "-C", script_dir, "diff", "--quiet", "data.json"],
            capture_output=True, creationflags=NO_WIN, timeout=10
        )
        if r.returncode != 0:
            run_git(["git", "-C", script_dir, "add", "data.json"])
            run_git(["git", "-C", script_dir, "commit", "-m",
                     f"auto: update data [{time.strftime('%H:%M')}]"])
            run_git(["git", "-C", script_dir, "push", "origin", "master"])
    except Exception:
        pass

# ── Main loop ──────────────────────────────────────────────────────────────────
# Load workbook sekali saat startup (butuh ~64 detik, hanya dilakukan satu kali)
try:
    get_wb()
except Exception:
    pass

while True:
    try:
        do_sync()
    except Exception:
        pass

    try:
        do_export()
    except Exception:
        pass

    try:
        push_if_changed()
    except Exception:
        pass

    time.sleep(INTERVAL)
