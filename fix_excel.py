"""
fix_excel.py — perbaiki Financial Management.xlsx:
  1. Conditional formatting: warna opaque + range B:U (termasuk kolom NTD)
  2. SUMIF formula di setiap sheet bulan (IDR row 4, NTD row 6)
  3. Garis tabel kolom U (NTD) sesuai style tabel
  4. Pie chart NTD per sheet bulan (anchor AI3)
  5. Main Dashboard: link akumulasi semua bulan
  6. Re-apply data validations
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, Reference

EXCEL_PATH = r"C:\Claude\Project\Financial\Financial Management.xlsx"

MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

# ── Warna baris per tipe (ARGB: FF = opaque) ──────────────────────────────────
TYPE_FILLS = {
    "Income":  PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    "Expanse": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
    "Invest":  PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
    "Savings": PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid"),
}

# Range sesuai tabel: B(date) C:E(type) F:H(cat) I:J(amount IDR) K:T(desc) U(amount NTD)
CF_RANGE = "B10:U1000"

# Kategori semua tipe (untuk dropdown)
ALL_CATEGORIES = (
    "Gaji,Bonus,Hasil Bisnis,Cashback,"
    "Makan & Minum,Transportasi,Tagihan,Belanja Bulanan,"
    "Internet,iCloud+,Netflix,Claude AI,"
    "Reksa Dana,Saham,Emas,Deposito,"
    "Dana Darurat,Tabungan Liburan,Tabungan Barang"
)

NTD_FMT = '_-"NT$ "* #,##0.00_-;\\-"NT$ "* #,##0.00_-;_-"NT$ "* "-"??_-;_-@_-'

_thin = Side(style="thin")
CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

print("Membuka file Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH)

# ── Fix setiap sheet bulan ─────────────────────────────────────────────────────
for month in MONTHS:
    if month not in wb.sheetnames:
        print(f"  [skip] sheet '{month}' tidak ditemukan")
        continue

    ws = wb[month]
    print(f"  [{month}] fix CF + border + SUMIF + chart NTD + validasi...")

    # 1. Hapus CF lama, apply ulang dengan warna benar dan range B:U
    ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
    for type_name, fill in TYPE_FILLS.items():
        ws.conditional_formatting.add(
            CF_RANGE,
            FormulaRule(formula=[f'$C10="{type_name}"'], fill=fill)
        )

    # 2. SUMIF IDR (kolom I) — row 4
    ws["E4"] = '=SUMIF($C$10:$C$1000,"Income",$I$10:$I$1000)'
    ws["I4"] = '=SUMIF($C$10:$C$1000,"Expanse",$I$10:$I$1000)'
    ws["M4"] = '=SUMIF($C$10:$C$1000,"Invest",$I$10:$I$1000)'
    ws["Q4"] = '=SUMIF($C$10:$C$1000,"Savings",$I$10:$I$1000)'

    # SUMIF NTD (kolom U = col 21) — row 6
    ws["E6"] = '=SUMIF($C$10:$C$1000,"Income",$U$10:$U$1000)'
    ws["I6"] = '=SUMIF($C$10:$C$1000,"Expanse",$U$10:$U$1000)'
    ws["M6"] = '=SUMIF($C$10:$C$1000,"Invest",$U$10:$U$1000)'
    ws["Q6"] = '=SUMIF($C$10:$C$1000,"Savings",$U$10:$U$1000)'

    # Revert formula TEXT di col I yang error → kembalikan ke =$E$8
    for r in range(11, 1001):
        cell_i = ws.cell(row=r, column=9)
        val = str(cell_i.value or '')
        if val.startswith('=IF(U') and 'TEXT' in val:
            cell_i.value = '=$E$8'

    # Hapus label yang salah posisi di row 9 col U (area banner)
    ws.cell(row=9, column=21).value = None

    # Header AMOUNT NT$ di row 10 (baris judul tabel)
    ws.cell(row=10, column=21).value = "AMOUNT NT$"

    # 3. Border + format NT$ kolom U rows 10–1000
    for r in range(10, 1001):
        cell = ws.cell(row=r, column=21)
        cell.border = CELL_BORDER
        if r >= 11:
            cell.number_format = NTD_FMT

    # 4. NTD chart helper data di col AD:AE (30:31) rows 3–7
    ws.cell(row=3, column=30).value = "Category"
    ws.cell(row=3, column=31).value = "NTD"
    ws.cell(row=4, column=30).value = "Income"
    ws.cell(row=4, column=31).value = "=E6"
    ws.cell(row=5, column=30).value = "Expanse"
    ws.cell(row=5, column=31).value = "=I6"
    ws.cell(row=6, column=30).value = "Invest"
    ws.cell(row=6, column=31).value = "=M6"
    ws.cell(row=7, column=30).value = "Savings"
    ws.cell(row=7, column=31).value = "=Q6"

    # Pie chart NTD — anchor AI3
    chart = PieChart()
    chart.title = f"NTD — {month}"
    chart.style = 10
    data = Reference(ws, min_col=31, min_row=3, max_row=7)
    cats = Reference(ws, min_col=30, min_row=4, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 14
    ws.add_chart(chart, "AI3")

    # 5. Re-apply data validations
    ws.data_validations.dataValidation = []

    # TYPE TRANSACTION dropdown — col C
    dv_type = DataValidation(
        type="list",
        formula1="Refference!$B$2:$E$2",
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.sqref = "C10:C1000"
    ws.add_data_validation(dv_type)

    # CATEGORY dropdown — col F
    dv_cat = DataValidation(
        type="list",
        formula1=f'"{ALL_CATEGORIES}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_cat.sqref = "F10:F1000"
    ws.add_data_validation(dv_cat)

# ── Fix Main Dashboard — akumulasi semua bulan ────────────────────────────────
print("  [Main Dashboard] link semua bulan...")
ws_main = wb["Main Dashboard"]

for i, month in enumerate(MONTHS):
    row = 9 + i  # Jan=9, Feb=10, ..., Dec=20
    # IDR
    ws_main.cell(row=row, column=4).value  = f"={month}!E4"   # D = Income IDR
    ws_main.cell(row=row, column=8).value  = f"={month}!I4"   # H = Expanse IDR
    ws_main.cell(row=row, column=12).value = f"={month}!M4"   # L = Invest IDR
    ws_main.cell(row=row, column=16).value = f"={month}!Q4"   # P = Savings IDR
    # NTD — kolom S,T,U,V (19,20,21,22)
    ws_main.cell(row=row, column=19).value = f"={month}!E6"   # S = Income NTD
    ws_main.cell(row=row, column=20).value = f"={month}!I6"   # T = Expanse NTD
    ws_main.cell(row=row, column=21).value = f"={month}!M6"   # U = Invest NTD
    ws_main.cell(row=row, column=22).value = f"={month}!Q6"   # V = Savings NTD

# Total row 21 — IDR
ws_main["D21"] = "=SUM(D9:D20)"
ws_main["H21"] = "=SUM(H9:H20)"
ws_main["L21"] = "=SUM(L9:L20)"
ws_main["P21"] = "=SUM(P9:P20)"

# Total row 21 — NTD
ws_main.cell(row=21, column=19).value = "=SUM(S9:S20)"
ws_main.cell(row=21, column=20).value = "=SUM(T9:T20)"
ws_main.cell(row=21, column=21).value = "=SUM(U9:U20)"
ws_main.cell(row=21, column=22).value = "=SUM(V9:V20)"

# Summary cards IDR row 5
ws_main["C5"] = "=D21"
ws_main["G5"] = "=H21"
ws_main["K5"] = "=L21"
ws_main["O5"] = "=P21"

# Header NTD — S8:V8 adalah merged cell, tulis hanya ke anchor S7
ws_main.cell(row=7, column=19).value = "NTD SUMMARY"
# S8:V8 merged → tidak bisa tulis ke T8/U8/V8 (secondary cells)

# Summary cards NTD row 5
ws_main.cell(row=5, column=19).value = "=S21"
ws_main.cell(row=5, column=20).value = "=T21"
ws_main.cell(row=5, column=21).value = "=U21"
ws_main.cell(row=5, column=22).value = "=V21"

# NTD pie chart helper data (vertikal) di col AH:AI (34:35) rows 3–7
ws_main.cell(row=3, column=34).value = "Category"
ws_main.cell(row=3, column=35).value = "NTD Total"
ws_main.cell(row=4, column=34).value = "Income"
ws_main.cell(row=4, column=35).value = "=S21"
ws_main.cell(row=5, column=34).value = "Expanse"
ws_main.cell(row=5, column=35).value = "=T21"
ws_main.cell(row=6, column=34).value = "Invest"
ws_main.cell(row=6, column=35).value = "=U21"
ws_main.cell(row=7, column=34).value = "Savings"
ws_main.cell(row=7, column=35).value = "=V21"

# NTD pie chart di Main Dashboard — anchor AI3
ws_ntd_chart = PieChart()
ws_ntd_chart.title = "NTD — All Months"
ws_ntd_chart.style = 10
data_main = Reference(ws_main, min_col=35, min_row=3, max_row=7)
cats_main = Reference(ws_main, min_col=34, min_row=4, max_row=7)
ws_ntd_chart.add_data(data_main, titles_from_data=True)
ws_ntd_chart.set_categories(cats_main)
ws_ntd_chart.width = 14
ws_ntd_chart.height = 14
ws_main.add_chart(ws_ntd_chart, "AI3")

# ── Save ──────────────────────────────────────────────────────────────────────
print("Menyimpan...")
try:
    wb.save(EXCEL_PATH)
    print("Selesai! File tersimpan.")
except PermissionError:
    print("ERROR: File Excel sedang terbuka. Tutup Excel dulu lalu jalankan ulang.")
