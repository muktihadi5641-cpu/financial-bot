"""
Financial Dashboard Server
Run: python financial_server.py
Access from PC:    http://localhost:5050
Access from phone: http://<your-PC-IP>:5050  (same WiFi)
"""
try:
    from flask import Flask, jsonify, send_file
except ImportError:
    print("Flask belum terinstall. Jalankan: pip install flask")
    input("Press Enter to exit...")
    exit(1)

import openpyxl
from datetime import datetime, date as date_type
import os, socket, logging, time

app = Flask(__name__)

EXCEL_PATH = r"C:\Claude\Project\Financial\Financial Management.xlsx"
MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
TYPE_MAP = {
    "Income":  "income",
    "Expanse": "expense",
    "Invest":  "invest",
    "Savings": "savings",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
log = logging.getLogger(__name__)

_cache = {"data": None, "ts": 0}
CACHE_TTL = 20  # seconds


def safe_num(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def fmt_date(v):
    if isinstance(v, (datetime, date_type)):
        d = v if isinstance(v, date_type) else v.date()
        return d.strftime("%d/%m/%Y")
    return str(v) if v else ""


def read_data():
    # read_only=True uses streaming XML parser — ~2000x faster than full load
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    months_data = {}

    for month in MONTHS:
        if month not in wb.sheetnames:
            continue
        ws = wb[month]
        transactions = []

        # min_col=1 so row[n] = column n+1 (A=0, B=1, C=2, F=5, I=8, K=10, U=20)
        for row in ws.iter_rows(min_row=11, max_row=1000, min_col=1, max_col=21, values_only=True):
            b_val = row[1]   # col B = date
            if b_val is None:
                continue     # skip sparse empty rows
            if not any(row[1:]):
                continue     # fully empty row
            transactions.append({
                "date":        fmt_date(b_val),
                "type":        row[2]  or "",   # col C
                "category":    row[5]  or "",   # col F
                "idr":         safe_num(row[8]),  # col I
                "ntd":         safe_num(row[20]), # col U
                "description": row[10] or "",   # col K
            })

        # Calculate summaries directly from transaction data (no formula caching issues)
        idr = {"income": 0.0, "expense": 0.0, "invest": 0.0, "savings": 0.0}
        ntd = {"income": 0.0, "expense": 0.0, "invest": 0.0, "savings": 0.0}
        for tx in transactions:
            k = TYPE_MAP.get(tx["type"])
            if k:
                idr[k] += tx["idr"]
                ntd[k] += tx["ntd"]

        months_data[month] = {
            "idr": idr,
            "ntd": ntd,
            "tx_count": len(transactions),
            "transactions": list(reversed(transactions)),  # newest first
        }

    wb.close()

    # Last modified time of Excel file
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
        last_sync = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
    except OSError:
        last_sync = "—"

    return {
        "months":    months_data,
        "last_sync": last_sync,
        "server_time": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }


def get_data_cached():
    now = time.time()
    if _cache["data"] is None or (now - _cache["ts"]) > CACHE_TTL:
        _cache["data"] = read_data()
        _cache["ts"] = now
    return _cache["data"]


@app.route("/")
def index():
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")
    resp = send_file(html_path, max_age=0)
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp


@app.route("/api/data")
def api_data():
    try:
        return jsonify({"ok": True, "data": get_data_cached()})
    except PermissionError:
        return jsonify({"ok": False, "error": "File Excel sedang terbuka di Excel. Tutup dulu."}), 503
    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"File tidak ditemukan: {EXCEL_PATH}"}), 404
    except Exception as e:
        log.error("Error reading Excel: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.after_request
def add_cors(r):
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


if __name__ == "__main__":
    try:
        local_ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        local_ip = "localhost"

    print("\n" + "=" * 50)
    print("  Financial Dashboard Server")
    print("=" * 50)
    print(f"  PC:    http://localhost:5050")
    print(f"  Phone: http://{local_ip}:5050")
    print("=" * 50)
    print("  Ctrl+C to stop\n")

    app.run(host="0.0.0.0", port=5050, debug=False)
