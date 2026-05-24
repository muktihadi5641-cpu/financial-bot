"""
Local Sync Script — Financial Management
Jalankan di Windows saat PC menyala (via Task Scheduler).
Mengambil transaksi pending dari bot cloud → tulis ke Excel.

Setup Task Scheduler:
  Trigger  : At startup + setiap 10 menit
  Action   : python "C:\\Claude\\Project\\Financial\\sync_to_excel.py"
  Start in : C:\\Claude\\Project\\Financial
"""

import os
import sys
import ssl
import json
import logging
import urllib.request
import urllib.error
from datetime import datetime

# Bypass SSL proxy jaringan lokal (sama seperti telegram_bot.py lokal)
_SSL_CTX = ssl._create_unverified_context()
from dotenv import load_dotenv
import openpyxl

load_dotenv()

RAILWAY_URL = os.getenv("RAILWAY_URL", "").rstrip("/")   # contoh: https://mybot.up.railway.app
SYNC_SECRET = os.getenv("SYNC_SECRET", "")
EXCEL_PATH  = os.getenv("EXCEL_PATH", r"C:\Claude\Project\Financial\Financial Management.xlsx")

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("sync.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

MONTH_SHEETS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]

# ── HTTP helper ───────────────────────────────────────────────────────────────

def api_get(path: str) -> list | dict | None:
    url = f"{RAILWAY_URL}{path}?token={SYNC_SECRET}"
    try:
        with urllib.request.urlopen(url, timeout=15, context=_SSL_CTX) as r:
            return json.loads(r.read())
    except urllib.error.URLError as e:
        log.error("Tidak bisa terhubung ke bot cloud: %s", e)
        return None

def api_post(path: str, data: dict) -> dict | None:
    url  = f"{RAILWAY_URL}{path}?token={SYNC_SECRET}"
    body = json.dumps(data).encode()
    req  = urllib.request.Request(url, data=body,
                                  headers={"Content-Type": "application/json"},
                                  method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15, context=_SSL_CTX) as r:
            return json.loads(r.read())
    except urllib.error.URLError as e:
        log.error("Gagal POST ke bot cloud: %s", e)
        return None

# ── Excel writer ──────────────────────────────────────────────────────────────

def find_next_row(ws) -> int | None:
    for row in range(10, 1001):
        if ws.cell(row=row, column=2).value is None:
            return row
    return None

def write_transactions(transactions: list[dict]) -> tuple[int, list[int]]:
    """Tulis list transaksi ke Excel. Return (jumlah_berhasil, list_id_berhasil)."""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except PermissionError:
        log.error("File Excel terbuka — tutup Excel lalu jalankan sync lagi.")
        return 0, []
    except FileNotFoundError:
        log.error("File tidak ditemukan: %s", EXCEL_PATH)
        return 0, []

    written  = 0
    done_ids = []

    for tx in transactions:
        date_str    = tx["date"]        # "2026-05-25"
        trans_type  = tx["type"]
        category    = tx["category"]
        amount      = float(tx["amount"])
        currency    = tx.get("currency", "IDR")   # "IDR" atau "NTD"
        description = tx["description"]

        # Tentukan sheet dari tanggal transaksi
        try:
            date_obj   = datetime.strptime(date_str, "%Y-%m-%d").date()
            sheet_name = MONTH_SHEETS[date_obj.month - 1]
        except Exception:
            log.warning("Tanggal tidak valid: %s — skip", date_str)
            continue

        if sheet_name not in wb.sheetnames:
            log.warning("Sheet '%s' tidak ada di workbook — skip", sheet_name)
            continue

        ws  = wb[sheet_name]
        row = find_next_row(ws)
        if row is None:
            log.warning("Sheet '%s' penuh — skip", sheet_name)
            continue

        # IDR → kolom I (9), NTD → kolom U (21, dipakai SUMIF NTD)
        col_amount = 9 if currency == "IDR" else 21

        ws.cell(row=row, column=2,          value=date_obj)
        ws.cell(row=row, column=3,          value=trans_type)
        ws.cell(row=row, column=6,          value=category)
        ws.cell(row=row, column=col_amount, value=amount)
        ws.cell(row=row, column=11,         value=description)
        ws.cell(row=row, column=2).number_format = "DD/MM/YYYY"

        # NTD: tulis formula display di col I agar tampil "NT$ X" bukan "Rp -"
        # SUMIF IDR tetap aman karena TEXT() menghasilkan string (diabaikan SUMIF numerik)
        if currency == "NTD":
            ws.cell(row=row, column=9).value = f'=IF(U{row}>0,TEXT(U{row},"NT$ #,##0"),$E$8)'

        curr_label = "NT$" if currency == "NTD" else "Rp"
        log.info("✓ [%s] %s | %s | %s %.0f | %s (baris %d)",
                 sheet_name, trans_type, category, curr_label, amount, description, row)
        written  += 1
        done_ids.append(tx["id"])

    if written > 0:
        try:
            wb.save(EXCEL_PATH)
            log.info("Excel tersimpan — %d transaksi ditulis.", written)
        except PermissionError:
            log.error("Gagal save — tutup Excel dulu.")
            return 0, []
    else:
        log.info("Tidak ada transaksi baru yang ditulis.")

    return written, done_ids

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not RAILWAY_URL:
        log.error("RAILWAY_URL belum diisi di .env")
        sys.exit(1)
    if not SYNC_SECRET:
        log.error("SYNC_SECRET belum diisi di .env")
        sys.exit(1)

    log.info("=== Sync dimulai: %s ===", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # 1. Ambil transaksi pending dari cloud
    pending = api_get("/pending")
    if pending is None:
        log.error("Gagal mengambil data dari cloud — cek koneksi internet.")
        sys.exit(1)

    if not pending:
        log.info("Tidak ada transaksi pending.")
        return

    log.info("Ditemukan %d transaksi pending.", len(pending))

    # 2. Tulis ke Excel
    count, done_ids = write_transactions(pending)

    # 3. Tandai sebagai sudah disync
    if done_ids:
        result = api_post("/mark_synced", {"ids": done_ids})
        if result:
            log.info("Ditandai synced: %d transaksi.", result.get("marked", 0))

    log.info("=== Sync selesai: %d/%d berhasil ===", count, len(pending))

if __name__ == "__main__":
    main()
