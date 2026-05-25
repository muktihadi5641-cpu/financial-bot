"""
copy_may_config.py — Copy konfigurasi sheet May (yang sudah sempurna) ke semua sheet bulan lain,
dan perbaiki referensi bulan di Main Dashboard agar sesuai.

Jalankan SEKALI dengan Excel tertutup.
"""

import openpyxl
from copy import copy
from openpyxl.cell.cell import MergedCell

EXCEL_PATH = r"C:\Claude\Project\Financial\Financial Management.xlsx"

MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]

# Nama bulan yang tampil di B3 (sesuai format yang terlihat di May = uppercase)
MONTH_UPPER = [m.upper() for m in MONTHS]

print("Membuka workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws_may = wb["May"]

# ── 1. Ambil dashboard merges dari May (baris 1–9) ────────────────────────────
print("Mengambil konfigurasi May...")

dashboard_merges = []
for mr in ws_may.merged_cells.ranges:
    if mr.min_row <= 9:
        dashboard_merges.append(str(mr))

# ── 2. Ambil values + styles dari May (baris 1–10) ───────────────────────────
dashboard_data = {}  # (row, col) → dict style
for row in ws_may.iter_rows(min_row=1, max_row=10, min_col=1, max_col=22):
    for cell in row:
        if isinstance(cell, MergedCell):
            continue  # skip secondary cells
        dashboard_data[(cell.row, cell.column)] = {
            "value":         cell.value,
            "font":          copy(cell.font),
            "fill":          copy(cell.fill),
            "border":        copy(cell.border),
            "alignment":     copy(cell.alignment),
            "number_format": cell.number_format,
        }

# ── 3. Ambil kolom lebar + tinggi baris dari May ─────────────────────────────
col_widths = {col: dim.width for col, dim in ws_may.column_dimensions.items() if dim.width}
row_heights = {
    int(rn): dim.height
    for rn, dim in ws_may.row_dimensions.items()
    if dim.height and int(rn) <= 10
}

# ── 4. Copy ke setiap sheet bulan (kecuali May) ───────────────────────────────
for idx, month in enumerate(MONTHS):
    if month == "May":
        print(f"  [May] skip — sudah sempurna")
        continue

    ws = wb[month]
    print(f"  [{month}] copy dashboard...")

    # Hapus merge lama di baris 1–9
    to_remove = [str(mr) for mr in ws.merged_cells.ranges if mr.min_row <= 9]
    for mr_str in to_remove:
        ws.merged_cells.remove(mr_str)

    # Apply merge dari May
    for mr_str in dashboard_merges:
        try:
            ws.merge_cells(mr_str)
        except Exception as e:
            print(f"    [warn] merge {mr_str}: {e}")

    # Copy cell values dan styles
    for (row, col), data in dashboard_data.items():
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue

        # B3 = nama bulan sesuai sheet ini
        if row == 3 and col == 2:
            cell.value = MONTH_UPPER[idx]
        else:
            cell.value = data["value"]

        cell.font          = data["font"]
        cell.fill          = data["fill"]
        cell.border        = data["border"]
        cell.alignment     = data["alignment"]
        cell.number_format = data["number_format"]

    # Copy lebar kolom dari May
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Copy tinggi baris 1–10 dari May
    for rn, height in row_heights.items():
        ws.row_dimensions[rn].height = height

# ── 5. Fix Main Dashboard — referensi bulan IDR dan NTD ──────────────────────
print("  [Main Dashboard] memperbaiki referensi bulan...")
ws_main = wb["Main Dashboard"]

# Rows 10–21 = Jan–Dec, row 22 = Total
for idx, month in enumerate(MONTHS):
    row = 10 + idx   # Jan=10, Feb=11, ..., Dec=21

    # IDR: D(4) Income, H(8) Expanse, L(12) Invest, P(16) Savings
    ws_main.cell(row=row, column=4).value  = f"={month}!E4"
    ws_main.cell(row=row, column=8).value  = f"={month}!I4"
    ws_main.cell(row=row, column=12).value = f"={month}!M4"
    ws_main.cell(row=row, column=16).value = f"={month}!Q4"

    # NTD: S(19) Income, T(20) Expanse, U(21) Invest, V(22) Savings
    ws_main.cell(row=row, column=19).value = f"={month}!E6"
    ws_main.cell(row=row, column=20).value = f"={month}!I6"
    ws_main.cell(row=row, column=21).value = f"={month}!M6"
    ws_main.cell(row=row, column=22).value = f"={month}!Q6"

# Baris label C10:C21 = Jan–Dec
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for idx, abbr in enumerate(MONTH_ABBR):
    ws_main.cell(row=10 + idx, column=3).value = abbr

# Total row 22 — IDR
ws_main["D22"] = "=SUM(D10:D21)"
ws_main["H22"] = "=SUM(H10:H21)"
ws_main["L22"] = "=SUM(L10:L21)"
ws_main["P22"] = "=SUM(P10:P21)"

# Total row 22 — NTD
ws_main.cell(row=22, column=19).value = "=SUM(S10:S21)"
ws_main.cell(row=22, column=20).value = "=SUM(T10:T21)"
ws_main.cell(row=22, column=21).value = "=SUM(U10:U21)"
ws_main.cell(row=22, column=22).value = "=SUM(V10:V21)"

# Label total
ws_main.cell(row=22, column=3).value = "Total"

# Summary cards IDR → sekarang referensi row 22 (total)
ws_main["C5"] = "=D22"
ws_main["G5"] = "=H22"
ws_main["K5"] = "=L22"
ws_main["O5"] = "=P22"

# Summary cards NTD → referensi row 22 (total)
ws_main.cell(row=7, column=3).value  = "=S22"
ws_main.cell(row=7, column=7).value  = "=T22"
ws_main.cell(row=7, column=11).value = "=U22"
ws_main.cell(row=7, column=15).value = "=V22"

# Hapus sisa data lama di baris 9 IDR (fix_excel.py sebelumnya menulis row 9)
for col in (4, 8, 12, 16):
    cell = ws_main.cell(row=9, column=col)
    if not isinstance(cell, MergedCell):
        # Kalau row 9 sudah berisi label "Spend" dari user, pertahankan
        # Kalau berisi formula lama (dari fix_excel.py), hapus
        val = str(cell.value or "")
        if val.startswith("="):
            cell.value = None

# ── 6. Simpan ─────────────────────────────────────────────────────────────────
print("Menyimpan...")
try:
    wb.save(EXCEL_PATH)
    print("Selesai! Semua sheet sudah disinkronkan dari May.")
except PermissionError:
    print("ERROR: File Excel sedang terbuka. Tutup Excel dulu lalu jalankan ulang.")
