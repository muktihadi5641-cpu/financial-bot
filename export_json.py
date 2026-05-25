"""
export_json.py — Export Financial Management.xlsx → data.json (untuk GitHub Pages dashboard).
Dijalankan otomatis oleh sync_loop.pyw setelah setiap sync.
"""
import openpyxl, json, os
from datetime import datetime, date as date_type

EXCEL_PATH = r"C:\Claude\Project\Financial\Financial Management.xlsx"
JSON_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
TYPE_MAP = {
    "Income":  "income",
    "Expanse": "expense",
    "Invest":  "invest",
    "Savings": "savings",
}


def safe_num(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def fmt_date(v):
    if isinstance(v, (datetime, date_type)):
        d = v if isinstance(v, date_type) else v.date()
        return d.strftime("%d/%m/%Y")
    return str(v) if v else ""


def export() -> bool:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    months_data = {}

    for month in MONTHS:
        if month not in wb.sheetnames:
            continue
        ws = wb[month]
        txs = []
        for row in ws.iter_rows(min_row=11, max_row=1000, min_col=1, max_col=21, values_only=True):
            b = row[1]
            if b is None:
                continue
            txs.append({
                "date":        fmt_date(b),
                "type":        row[2]  or "",
                "category":    row[5]  or "",
                "idr":         safe_num(row[8]),
                "ntd":         safe_num(row[20]),
                "description": row[10] or "",
            })

        idr = {"income": 0.0, "expense": 0.0, "invest": 0.0, "savings": 0.0}
        ntd = {"income": 0.0, "expense": 0.0, "invest": 0.0, "savings": 0.0}
        for tx in txs:
            k = TYPE_MAP.get(tx["type"])
            if k:
                idr[k] += tx["idr"]
                ntd[k] += tx["ntd"]

        months_data[month] = {
            "idr":          idr,
            "ntd":          ntd,
            "tx_count":     len(txs),
            "transactions": list(reversed(txs)),
        }

    wb.close()

    try:
        mtime     = os.path.getmtime(EXCEL_PATH)
        last_sync = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
    except OSError:
        last_sync = "—"

    payload = {
        "ok": True,
        "data": {
            "months":    months_data,
            "last_sync": last_sync,
        }
    }

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    tx_total = sum(m["tx_count"] for m in months_data.values())
    print(f"[export_json] {len(months_data)} months, {tx_total} tx -> data.json")
    return True


if __name__ == "__main__":
    export()
