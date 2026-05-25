"""
deep_clone_may.py — Kloning LENGKAP sheet May ke semua sheet bulan lain.

Apa yang dicopy:
  - Semua lebar kolom (B=17, U=12.89, AC=9.33, kolom lain reset ke default)
  - Tinggi baris: row 2=25.2, row 6=21.0, row 11-1001=15.6, default=14.4
  - Merged cells baris 1-10 (dashboard + header tabel)
  - Nilai & style semua sel baris 1-10 (B3 diganti nama bulan)
  - Helper data IDR di AB2:AC5, NTD di AB20:AC23
  - 2 DoughnutChart: IDR di W1 (15x7.5cm), NTD di W18 (15x7.5cm)
  - Conditional formatting (Income/Expanse/Invest/Savings → warna baris)
  - Data validation (dropdown Type & Category)
  - SUMIF formulas di row 4 (IDR) dan row 6 (NTD)
  - NTD format NT$ di kolom U baris 11-1000

Jalankan SEKALI dengan Excel tertutup.
"""

import openpyxl
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.styles import PatternFill

EXCEL_PATH = r"C:\Claude\Project\Financial\Financial Management.xlsx"

MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
MONTH_UPPER = [m.upper() for m in MONTHS]

# Conditional formatting: warna baris berdasarkan tipe transaksi
TYPE_FILLS = {
    "Income":  PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    "Expanse": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
    "Invest":  PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
    "Savings": PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid"),
}
CF_RANGE = "B10:U1000"

ALL_CATEGORIES = (
    "Gaji,Bonus,Hasil Bisnis,Cashback,"
    "Makan & Minum,Transportasi,Tagihan,Belanja Bulanan,"
    "Internet,iCloud+,Netflix,Claude AI,"
    "Reksa Dana,Saham,Emas,Deposito,"
    "Dana Darurat,Tabungan Liburan,Tabungan Barang"
)
NTD_FMT = '_-"NT$ "* #,##0.00_-;\\-"NT$ "* #,##0.00_-;_-"NT$ "* "-"??_-;_-@_-'

# ── Row heights May (dari inspeksi aktual) ─────────────────────────────────────
# Template rows dengan tinggi khusus
TEMPLATE_ROW_HEIGHTS = {2: 25.2, 6: 21.0, 1002: 18.0}
# Rows 3,4,5,7,8,9,10 = 14.4 (default) — tidak perlu set eksplisit
# Rows 11-1001 = 15.6 (data rows)
DATA_ROW_HEIGHT = 15.6
DEFAULT_ROW_HEIGHT = 14.4

print("Membuka workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws_may = wb["May"]

# ── Baca semua data dari May ───────────────────────────────────────────────────
print("Membaca properti May...")

# Cell values + styles untuk baris 1-10, kolom A-U (1-21)
may_cells = {}
for row in ws_may.iter_rows(min_row=1, max_row=10, min_col=1, max_col=21):
    for cell in row:
        if isinstance(cell, MergedCell):
            continue
        may_cells[(cell.row, cell.column)] = {
            "value":         cell.value,
            "font":          copy(cell.font),
            "fill":          copy(cell.fill),
            "border":        copy(cell.border),
            "alignment":     copy(cell.alignment),
            "number_format": cell.number_format,
        }

# Merged cells baris 1-10 dari May
dashboard_merges = [str(mr) for mr in ws_may.merged_cells.ranges if mr.min_row <= 10]

print(f"  {len(may_cells)} sel, {len(dashboard_merges)} merged ranges (baris 1-10)")
print(f"  Charts di May: {len(ws_may._charts)}")

# ── Clone ke setiap sheet bulan (kecuali May) ─────────────────────────────────
for idx, month in enumerate(MONTHS):
    if month == "May":
        print(f"\n  [May] skip — sudah sempurna")
        continue

    ws = wb[month]
    print(f"\n  [{month}]")

    # ── 1. Lebar kolom ─────────────────────────────────────────────────────────
    # Semua bulan sudah punya B=17, U=12.89, AC=9.33 — pastikan sesuai May
    ws.column_dimensions["B"].width  = 17.0
    ws.column_dimensions["U"].width  = 12.88671875
    ws.column_dimensions["AC"].width = 9.33203125
    print(f"    OK column widths (B=17, U=12.89, AC=9.33)")

    # ── 2. Tinggi baris ────────────────────────────────────────────────────────
    # Default sheet row height
    ws.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT
    ws.sheet_format.customHeight = True

    # Reset baris 1-10 ke default dulu, lalu set sesuai May
    for rn in range(1, 11):
        ws.row_dimensions[rn].height = DEFAULT_ROW_HEIGHT
    # Template rows khusus
    for rn, h in TEMPLATE_ROW_HEIGHTS.items():
        ws.row_dimensions[rn].height = h
    # Baris data (11-1001) = 15.6 sesuai May
    for rn in range(11, 1002):
        ws.row_dimensions[rn].height = DATA_ROW_HEIGHT
    print(f"    OK row heights (rows 11-1001 = {DATA_ROW_HEIGHT}, default = {DEFAULT_ROW_HEIGHT})")

    # ── 3. Merged cells baris 1-10 ────────────────────────────────────────────
    # Hapus merges lama di baris 1-10
    to_remove = [str(mr) for mr in ws.merged_cells.ranges if mr.min_row <= 10]
    for mr_str in to_remove:
        ws.merged_cells.remove(mr_str)
    # Apply merges dari May
    warn_count = 0
    for mr_str in dashboard_merges:
        try:
            ws.merge_cells(mr_str)
        except Exception:
            warn_count += 1
    print(f"    OK merged cells ({len(dashboard_merges)} ranges, {warn_count} skip)")

    # ── 4. Nilai & style sel baris 1-10 ───────────────────────────────────────
    copied_count = 0
    for (row, col), data in may_cells.items():
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        # B3 = nama bulan yang benar
        if row == 3 and col == 2:
            cell.value = MONTH_UPPER[idx]
        else:
            cell.value = data["value"]
        cell.font          = copy(data["font"])
        cell.fill          = copy(data["fill"])
        cell.border        = copy(data["border"])
        cell.alignment     = copy(data["alignment"])
        cell.number_format = data["number_format"]
        copied_count += 1
    print(f"    OK cell values & styles ({copied_count} sel)")

    # ── 5. SUMIF formulas (IDR row 4, NTD row 6) ──────────────────────────────
    ws["E4"] = '=SUMIF($C$10:$C$1000,"Income",$I$10:$I$1000)'
    ws["I4"] = '=SUMIF($C$10:$C$1000,"Expanse",$I$10:$I$1000)'
    ws["M4"] = '=SUMIF($C$10:$C$1000,"Invest",$I$10:$I$1000)'
    ws["Q4"] = '=SUMIF($C$10:$C$1000,"Savings",$I$10:$I$1000)'
    ws["E6"] = '=SUMIF($C$10:$C$1000,"Income",$U$10:$U$1000)'
    ws["I6"] = '=SUMIF($C$10:$C$1000,"Expanse",$U$10:$U$1000)'
    ws["M6"] = '=SUMIF($C$10:$C$1000,"Invest",$U$10:$U$1000)'
    ws["Q6"] = '=SUMIF($C$10:$C$1000,"Savings",$U$10:$U$1000)'
    print(f"    OK SUMIF formulas (IDR row 4, NTD row 6)")

    # ── 6. NTD kolom U ─────────────────────────────────────────────────────────
    ws.cell(row=10, column=21).value = "AMOUNT NT$"
    for r in range(11, 1001):
        cell = ws.cell(row=r, column=21)
        if not isinstance(cell, MergedCell):
            cell.number_format = NTD_FMT
    print(f"    OK NTD column U (header + format NT$)")

    # ── 7. Helper data untuk chart ─────────────────────────────────────────────
    # IDR helper di AB2:AC5 (kolom 28:29)
    ws.cell(row=2, column=28).value = "Income";   ws.cell(row=2, column=29).value = "=E4"
    ws.cell(row=3, column=28).value = "Expanse";  ws.cell(row=3, column=29).value = "=I4"
    ws.cell(row=4, column=28).value = "Invest";   ws.cell(row=4, column=29).value = "=M4"
    ws.cell(row=5, column=28).value = "Savings";  ws.cell(row=5, column=29).value = "=Q4"
    # NTD helper di AB20:AC23 (kolom 28:29)
    ws.cell(row=20, column=28).value = "Income";  ws.cell(row=20, column=29).value = "=E6"
    ws.cell(row=21, column=28).value = "Expanse"; ws.cell(row=21, column=29).value = "=I6"
    ws.cell(row=22, column=28).value = "Invest";  ws.cell(row=22, column=29).value = "=M6"
    ws.cell(row=23, column=28).value = "Savings"; ws.cell(row=23, column=29).value = "=Q6"
    print(f"    OK chart helper data (AB2:AC5 IDR, AB20:AC23 NTD)")

    # ── 8. Charts: hapus lama, buat 2 DoughnutChart baru ──────────────────────
    ws._charts = []

    # IDR DoughnutChart di W1 — size 15x7.5 cm
    idr_chart = DoughnutChart()
    idr_chart.title = f"{month} - (IDR) Financial Distribution"
    idr_chart.style = 10
    idr_chart.width  = 15
    idr_chart.height = 7.5
    idr_data = Reference(ws, min_col=29, min_row=2, max_row=5)  # AC2:AC5
    idr_cats = Reference(ws, min_col=28, min_row=2, max_row=5)  # AB2:AB5
    idr_chart.add_data(idr_data)
    idr_chart.set_categories(idr_cats)
    ws.add_chart(idr_chart, "W1")

    # NTD DoughnutChart di W18 — size 15x7.5 cm
    ntd_chart = DoughnutChart()
    ntd_chart.title = f"{month} - (NTD) Financial Distribution"
    ntd_chart.style = 10
    ntd_chart.width  = 15
    ntd_chart.height = 7.5
    ntd_data = Reference(ws, min_col=29, min_row=20, max_row=23)  # AC20:AC23
    ntd_cats = Reference(ws, min_col=28, min_row=20, max_row=23)  # AB20:AB23
    ntd_chart.add_data(ntd_data)
    ntd_chart.set_categories(ntd_cats)
    ws.add_chart(ntd_chart, "W18")

    print(f"    OK 2 DoughnutCharts (IDR@W1, NTD@W18 — masing-masing 15x7.5cm)")

    # ── 9. Conditional formatting ──────────────────────────────────────────────
    ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
    for type_name, fill in TYPE_FILLS.items():
        ws.conditional_formatting.add(
            CF_RANGE,
            FormulaRule(formula=[f'$C10="{type_name}"'], fill=fill)
        )
    print(f"    OK conditional formatting (4 rules, range {CF_RANGE})")

    # ── 10. Data validations ───────────────────────────────────────────────────
    ws.data_validations.dataValidation = []
    dv_type = DataValidation(
        type="list",
        formula1="Refference!$B$2:$E$2",
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.sqref = "C10:C1000"
    ws.add_data_validation(dv_type)

    dv_cat = DataValidation(
        type="list",
        formula1=f'"{ALL_CATEGORIES}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_cat.sqref = "F10:F1000"
    ws.add_data_validation(dv_cat)
    print(f"    OK data validations (Type & Category dropdowns)")

# ── Simpan ─────────────────────────────────────────────────────────────────────
print("\nMenyimpan workbook...")
try:
    wb.save(EXCEL_PATH)
    print("=" * 55)
    print("SELESAI! Semua 11 sheet bulan sudah kloning dari May.")
    print("  - Layout dashboard identik (merged cells, styles, font)")
    print("  - Ukuran kolom B=17, U=12.89, AC=9.33")
    print("  - Tinggi baris data (11-1001) = 15.6, template = 14.4")
    print("  - 2 DoughnutChart per sheet: IDR@W1, NTD@W18")
    print("  - Conditional formatting + dropdowns aktif")
    print("=" * 55)
except PermissionError:
    print("\nERROR: File Excel sedang terbuka. Tutup dulu lalu jalankan ulang.")
